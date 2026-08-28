#!/usr/bin/env python3
"""
scripts/build-data.py — turn the supplied .xlsx into compact JSON the browser
can load and compute over.

WHY A BUILD STEP. The workbook is 4.8 MB of XML across fourteen sheets. Parsing
that in a browser needs a spreadsheet library and several seconds on a phone.
Converting once, here, keeps the running app dependency-free and fast — and,
more importantly, keeps the conversion inspectable: this file is the entire
description of what was taken from the workbook and what was left behind.

NOTHING IS AGGREGATED HERE. Every row of the fact table survives, so the app
computes its own totals from transaction level. A build step that pre-computed
the answers would be a build step that could hide a mistake.

Strings that repeat (regions, couriers, channels) are interned into lookup
tables and stored as integers. That is a storage decision, not a data decision:
the values round-trip exactly.

    python3 scripts/build-data.py
"""

import json
import os
import sys
from collections import OrderedDict

import openpyxl

SRC = os.environ.get(
    "NEXA_XLSX",
    os.path.expanduser("~/Downloads/NexaSphere_BI_Case_Study_Dataset.xlsx"),
)
OUT = os.path.join(os.path.dirname(__file__), "..", "data", "nexasphere.json")

# Columns kept from Fact_Sales. Everything needed to answer the nine management
# questions in the case brief, and nothing else.
SALES_COLS = [
    "Order_ID", "Order_Month", "Customer_ID", "Customer_Type",
    "Product_ID", "Store_ID", "Sales_Agent_ID", "Campaign_ID",
    "Sales_Channel", "Customer_Region",
    "Quantity", "Discount_Pct", "Net_Sales_NGN", "Discount_Amount_NGN",
    "Realized_Revenue_NGN", "COGS_Recognized_NGN", "Gross_Profit_NGN",
    "Contribution_Profit_NGN", "Refund_Amount_NGN",
    "Courier", "Promised_Days", "Delivery_Days", "Late_Delivery_Flag",
    "Order_Status", "Return_Flag", "Return_Reason",
    "Customer_Rating", "Support_Tickets", "Batch_Code",
]

# Text columns worth interning — low cardinality, high repetition.
INTERN = {
    "Customer_Type", "Sales_Channel", "Customer_Region", "Courier",
    "Order_Status", "Return_Reason", "Batch_Code", "Order_Month",
}


def jsonable(v):
    """Dates arrive as datetimes from openpyxl; JSON has no such type."""
    if hasattr(v, "isoformat"):
        return v.isoformat()[:10]
    return v


def sheet_rows(wb, name):
    ws = wb[name]
    it = ws.iter_rows(values_only=True)
    header = [str(h) if h is not None else "" for h in next(it)]
    for row in it:
        if row is None or all(v is None for v in row):
            continue
        yield header, row


def read_dim(wb, name):
    """A dimension table, as a list of dicts. These are small."""
    out = []
    header = None
    for header, row in sheet_rows(wb, name):
        out.append({h: jsonable(v) for h, v in zip(header, row) if h})
    return out


def money(v):
    """Naira to integer kobo. Floats never survive into the app."""
    if v is None:
        return 0
    return int(round(float(v) * 100))


def flag(v):
    return 1 if v in (1, True, "1", "Y", "Yes", "TRUE", "True") else 0


def num(v):
    if v is None:
        return 0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0


def main():
    if not os.path.exists(SRC):
        sys.exit(f"workbook not found: {SRC}\nSet NEXA_XLSX to its path.")

    print(f"  reading {SRC}")
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

    # ---- Fact_Sales ------------------------------------------------------
    ws = wb["Fact_Sales"]
    it = ws.iter_rows(values_only=True)
    header = [str(h) if h is not None else "" for h in next(it)]
    idx = {h: i for i, h in enumerate(header)}

    missing = [c for c in SALES_COLS if c not in idx]
    if missing:
        sys.exit(f"workbook is missing expected columns: {missing}")

    interns = {c: OrderedDict() for c in INTERN}

    def intern(col, value):
        table = interns[col]
        key = "" if value is None else str(value)
        if key not in table:
            table[key] = len(table)
        return table[key]

    MONEY_COLS = {
        "Net_Sales_NGN", "Discount_Amount_NGN", "Realized_Revenue_NGN",
        "COGS_Recognized_NGN", "Gross_Profit_NGN", "Contribution_Profit_NGN",
        "Refund_Amount_NGN",
    }
    FLAG_COLS = {"Late_Delivery_Flag", "Return_Flag"}

    rows = []
    for raw in it:
        if raw is None or raw[0] is None:
            continue
        rec = []
        for c in SALES_COLS:
            v = raw[idx[c]]
            if c in INTERN:
                rec.append(intern(c, v))
            elif c in MONEY_COLS:
                rec.append(money(v))
            elif c in FLAG_COLS:
                rec.append(flag(v))
            elif c in ("Quantity", "Promised_Days", "Delivery_Days",
                       "Support_Tickets", "Customer_Rating", "Discount_Pct"):
                rec.append(num(v))
            else:
                rec.append("" if v is None else str(jsonable(v)))
        rows.append(rec)

    print(f"  Fact_Sales: {len(rows):,} rows x {len(SALES_COLS)} columns")

    payload = {
        "generated": True,
        "source": os.path.basename(SRC),
        "sales": {
            "columns": SALES_COLS,
            "intern": {c: list(interns[c].keys()) for c in INTERN},
            "moneyColumns": sorted(MONEY_COLS),
            "rows": rows,
        },
        "products": read_dim(wb, "Dim_Products"),
        "stores": read_dim(wb, "Dim_Stores"),
        "employees": read_dim(wb, "Dim_Employees"),
        "campaigns": read_dim(wb, "Dim_Campaigns"),
        "inventory": read_dim(wb, "Fact_Inventory_Monthly"),
        "marketing": read_dim(wb, "Fact_Marketing"),
        "targets": read_dim(wb, "Targets_Monthly"),
    }

    for k in ("products", "stores", "employees", "campaigns", "inventory",
              "marketing", "targets"):
        print(f"  {k}: {len(payload[k]):,}")

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w") as fh:
        json.dump(payload, fh, separators=(",", ":"))

    size = os.path.getsize(OUT) / 1024 / 1024
    print(f"\n  wrote {OUT} ({size:.1f} MB)")


if __name__ == "__main__":
    main()
